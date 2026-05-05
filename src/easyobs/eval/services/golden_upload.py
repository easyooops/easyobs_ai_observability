"""Golden Set upload parser with security hardening.

Implements 11 §8.3 / 12 §9 — accepts CSV / JSONL / xlsx uploads, applies
the column mapping, and persists rows as ``candidate`` Golden Items
without touching the existing trace ingest path.

Security guards (mandatory — see 11 §8 / 12 §9.3):

1. **Formula injection** — every cell whose value starts with one of
   ``= + - @`` (and the leading-tab/CR variants) is prefixed with a
   single quote so spreadsheets cannot evaluate it on download.
2. **Macro rejection** — ``.xlsm`` / ``.xltm`` files are refused
   outright. The router only accepts ``.xlsx``.
3. **External link suppression** — ``openpyxl`` is opened in
   ``read_only=True`` + ``data_only=True`` mode, which never resolves
   external workbook references.
4. **Row / column / size caps** — enforced from
   :class:`easyobs.settings.Settings`.
5. **Optional PII redaction** — when the caller asks for it, every
   string cell is run through a fixed set of email / phone / Korean RRN
   patterns and masked.

The parser intentionally lives outside :mod:`easyobs.eval.services.goldensets`
so it does not pull ``openpyxl`` into the import graph for installs that
never touch upload (low-overhead deployments).
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Iterable, Sequence

_log = logging.getLogger("easyobs.eval.upload")


# ---------------------------------------------------------------------------
# File-kind detection / errors
# ---------------------------------------------------------------------------

# Files we accept. ``.xlsm``/``.xltm`` are deliberately absent — those
# carry VBA macros and would expand the attack surface.
ALLOWED_EXTENSIONS = {"csv", "jsonl", "xlsx"}


class UploadError(ValueError):
    """Raised on any caller-visible parser error.

    Carries an optional ``code`` so the router can map errors to a
    stable shape without string-matching messages.
    """

    def __init__(self, message: str, *, code: str = "upload_error") -> None:
        super().__init__(message)
        self.code = code


def detect_file_kind(filename: str) -> str:
    name = (filename or "").lower().rsplit(".", 1)
    if len(name) != 2:
        raise UploadError("missing file extension", code="invalid_extension")
    ext = name[1]
    if ext in {"xlsm", "xltm", "xltx"}:
        # Macros / external link templates rejected up front.
        raise UploadError(
            "macro-enabled spreadsheets are not allowed", code="macro_rejected"
        )
    if ext not in ALLOWED_EXTENSIONS:
        raise UploadError(
            f"unsupported file type: {ext}", code="invalid_extension"
        )
    return ext


# ---------------------------------------------------------------------------
# Cell sanitisation
# ---------------------------------------------------------------------------

# Characters that trigger formula evaluation in Excel / LibreOffice / Numbers.
# We also include the leading whitespace variants because a cell starting
# with ``\t=`` is still treated as a formula by some clients.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitise_cell(value: object) -> Any:
    """Return the safe form of a single cell value.

    - ``None`` / non-string scalars pass through unchanged.
    - Strings whose first character matches a formula prefix get a
      single quote prepended, neutralising the formula on re-export.
    """

    if value is None:
        return None
    if isinstance(value, bool):
        # Important: bool is a subclass of int — handle before int branch.
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text and text[0] in _FORMULA_PREFIXES:
        return "'" + text
    return text


# Light-touch PII patterns. Source order matters — phone before generic
# digits so we don't double-mask. Email is anchored on the ``@`` so we
# don't false-match inside JSON-encoded fields.
_PII_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I), "<email>"),
    # KR landline + mobile (010-1234-5678 / 02-123-4567 etc.)
    (re.compile(r"\b0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4}\b"), "<phone>"),
    # KR resident registration number — 6 digits + dash + 7 digits.
    (re.compile(r"\b\d{6}[-\s]?\d{7}\b"), "<rrn>"),
)


def _redact_pii(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        return value
    out = value
    for pat, mask in _PII_PATTERNS:
        out = pat.sub(mask, out)
    return out


# ---------------------------------------------------------------------------
# Mapping → Golden Item payload
# ---------------------------------------------------------------------------

# Closed list of fields the UI may map to. Anything else gets ignored
# — the mapping is operator-supplied, so we never trust it as keys.
SUPPORTED_GOLDEN_PATHS = (
    # L1 (query)
    "L1.query_text",
    "L1.intent",
    "L1.difficulty",
    "L1.language",
    "L1.expected_tool",
    "L1.tags",
    # L2 (retrieval) — list-shaped keys collect comma-split values.
    "L2.relevant_doc_ids",
    "L2.must_have_chunks",
    "L2.k_target",
    # L3 (response)
    "L3.expected_answer_text",
    "L3.must_include",
    "L3.must_not_include",
    "L3.citations_expected",
    "L3.schema",
)
_LIST_PATHS = {
    "L1.tags",
    "L2.relevant_doc_ids",
    "L2.must_have_chunks",
    "L3.must_include",
    "L3.must_not_include",
    "L3.citations_expected",
}


def _split_list(raw: Any) -> list[str]:
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    if raw is None:
        return []
    text = str(raw)
    if not text.strip():
        return []
    # Comma- or semicolon-delimited; strip the formula-injection
    # prefix we added so list elements stay readable.
    parts = re.split(r"[,;\n]", text)
    return [p.strip().lstrip("'") for p in parts if p.strip()]


def _payload_from_row(
    *, row: dict[str, Any], mapping: dict[str, str]
) -> dict[str, Any]:
    """Project a parsed row through ``mapping`` (csv_col -> golden_path)
    and return the layered payload accepted by ``goldensets`` service."""

    out: dict[str, Any] = {"L1": {}, "L2": {}, "L3": {}}
    for csv_col, golden_path in mapping.items():
        if golden_path not in SUPPORTED_GOLDEN_PATHS:
            continue
        layer, _, key = golden_path.partition(".")
        raw = row.get(csv_col)
        value: Any
        if golden_path in _LIST_PATHS:
            value = _split_list(raw)
        elif golden_path == "L2.k_target":
            try:
                value = int(str(raw).strip()) if raw not in (None, "") else None
            except ValueError:
                value = None
        elif golden_path == "L3.schema":
            # JSON schema is mapped verbatim — we only validate it parses.
            if isinstance(raw, str) and raw.strip():
                try:
                    value = json.loads(raw)
                except json.JSONDecodeError:
                    value = raw
            else:
                value = raw
        else:
            value = raw if raw is not None else ""
        out[layer][key] = value
    # Drop empty layers so the persisted payload stays compact.
    return {layer: data for layer, data in out.items() if data}


# ---------------------------------------------------------------------------
# Parsers per file kind
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class ParsedRow:
    index: int
    row: dict[str, Any]
    issues: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ParseResult:
    headers: list[str]
    rows: list[ParsedRow]
    skipped: int = 0
    truncated: bool = False
    file_kind: str = ""

    def total(self) -> int:
        return len(self.rows) + self.skipped


def _parse_csv(
    data: bytes,
    *,
    has_header: bool,
    max_rows: int,
    max_cols: int,
    redact: bool,
) -> ParseResult:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return ParseResult(headers=[], rows=[], file_kind="csv")
    if has_header:
        headers = [str(h).strip() or f"col_{i}" for i, h in enumerate(rows[0][:max_cols])]
        body = rows[1:]
    else:
        headers = [f"col_{i}" for i in range(min(len(rows[0]), max_cols))]
        body = rows
    return _normalise_rows(
        headers=headers,
        body=body,
        max_rows=max_rows,
        max_cols=max_cols,
        redact=redact,
        file_kind="csv",
    )


def _parse_jsonl(
    data: bytes, *, max_rows: int, max_cols: int, redact: bool
) -> ParseResult:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    parsed: list[ParsedRow] = []
    headers: list[str] = []
    seen: set[str] = set()
    skipped = 0
    truncated = False
    for idx, line in enumerate(text.splitlines()):
        if not line.strip():
            continue
        if len(parsed) >= max_rows:
            truncated = True
            break
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            skipped += 1
            continue
        if not isinstance(obj, dict):
            skipped += 1
            continue
        clean: dict[str, Any] = {}
        for key, value in list(obj.items())[:max_cols]:
            if key not in seen:
                seen.add(key)
                headers.append(str(key))
            sanitised = _sanitise_cell(value)
            if redact:
                sanitised = _redact_pii(sanitised)
            clean[str(key)] = sanitised
        parsed.append(ParsedRow(index=idx, row=clean))
    return ParseResult(
        headers=headers,
        rows=parsed,
        skipped=skipped,
        truncated=truncated,
        file_kind="jsonl",
    )


def _parse_xlsx(
    data: bytes,
    *,
    has_header: bool,
    max_rows: int,
    max_cols: int,
    redact: bool,
) -> ParseResult:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
        raise UploadError(
            "openpyxl is not installed on this server", code="missing_dependency"
        ) from exc
    try:
        wb = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise UploadError(
            f"unable to read xlsx workbook: {exc}", code="invalid_xlsx"
        ) from exc
    try:
        ws = wb.active
        if ws is None:
            return ParseResult(headers=[], rows=[], file_kind="xlsx")
        body: list[list[Any]] = []
        truncated = False
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx >= max_rows + (1 if has_header else 0):
                truncated = True
                break
            body.append(list(row[:max_cols]))
    finally:
        wb.close()

    if not body:
        return ParseResult(headers=[], rows=[], file_kind="xlsx", truncated=truncated)
    if has_header:
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(body[0])]
        rows = body[1:]
    else:
        headers = [f"col_{i}" for i in range(len(body[0]))]
        rows = body
    return _normalise_rows(
        headers=headers,
        body=rows,
        max_rows=max_rows,
        max_cols=max_cols,
        redact=redact,
        file_kind="xlsx",
        truncated=truncated,
    )


def _normalise_rows(
    *,
    headers: list[str],
    body: Iterable[Sequence[Any]],
    max_rows: int,
    max_cols: int,
    redact: bool,
    file_kind: str,
    truncated: bool = False,
) -> ParseResult:
    parsed: list[ParsedRow] = []
    skipped = 0
    for idx, raw in enumerate(body):
        if len(parsed) >= max_rows:
            truncated = True
            break
        if not raw or all(cell in (None, "") for cell in raw):
            skipped += 1
            continue
        clean: dict[str, Any] = {}
        for i, header in enumerate(headers[:max_cols]):
            value = raw[i] if i < len(raw) else None
            sanitised = _sanitise_cell(value)
            if redact:
                sanitised = _redact_pii(sanitised)
            clean[header] = sanitised
        parsed.append(ParsedRow(index=idx, row=clean))
    return ParseResult(
        headers=headers[:max_cols],
        rows=parsed,
        skipped=skipped,
        truncated=truncated,
        file_kind=file_kind,
    )


# ---------------------------------------------------------------------------
# Public entry-point
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class UploadValidation:
    """Outcome of :func:`validate_upload`. Mirrors the JSON shape the
    router returns to the UI's preview pane."""

    file_kind: str
    headers: list[str]
    sample_rows: list[dict[str, Any]]
    valid_count: int
    skipped_count: int
    truncated: bool
    issues: list[str]
    payloads: list[dict[str, Any]]


def validate_upload(
    *,
    filename: str,
    data: bytes,
    mapping: dict[str, str],
    has_header: bool,
    redact_pii: bool,
    max_size_bytes: int,
    max_rows: int,
    max_cols: int,
    sample_size: int = 50,
) -> UploadValidation:
    """Parse + sanitise + map an upload payload.

    The same function powers both the *preview* path (caller looks at
    ``sample_rows`` / ``issues``) and the *consume* path (caller writes
    ``payloads`` straight into Golden Items). Keeping a single entry
    point means the preview cannot diverge from the persisted result.
    """

    if max_size_bytes > 0 and len(data) > max_size_bytes:
        raise UploadError(
            f"file too large: {len(data)} bytes > {max_size_bytes}",
            code="size_exceeded",
        )
    file_kind = detect_file_kind(filename)
    if file_kind == "csv":
        result = _parse_csv(
            data,
            has_header=has_header,
            max_rows=max_rows,
            max_cols=max_cols,
            redact=redact_pii,
        )
    elif file_kind == "jsonl":
        result = _parse_jsonl(
            data, max_rows=max_rows, max_cols=max_cols, redact=redact_pii
        )
    elif file_kind == "xlsx":
        result = _parse_xlsx(
            data,
            has_header=has_header,
            max_rows=max_rows,
            max_cols=max_cols,
            redact=redact_pii,
        )
    else:  # pragma: no cover — detect_file_kind already validates
        raise UploadError(f"unsupported kind {file_kind}", code="invalid_extension")

    payloads: list[dict[str, Any]] = []
    issues: list[str] = []
    for parsed in result.rows:
        payload = _payload_from_row(row=parsed.row, mapping=mapping)
        if not payload:
            issues.append(f"row {parsed.index}: no mapped fields")
            continue
        payloads.append(payload)

    return UploadValidation(
        file_kind=result.file_kind,
        headers=result.headers,
        sample_rows=[r.row for r in result.rows[:sample_size]],
        valid_count=len(payloads),
        skipped_count=result.skipped,
        truncated=result.truncated,
        issues=issues[:50],
        payloads=payloads,
    )
